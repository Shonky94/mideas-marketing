import os
import pandas as pd
import re
import openpyxl
from openpyxl import load_workbook

def is_empty_column(df, col_name):
    """Check if a column is empty (all values are NaN or empty string)"""
    return df[col_name].isna().all() or (df[col_name] == '').all()

def contains_links(df, col_name):
    """Check if a column contains any links (http or www)"""
    if df[col_name].dtype == object:  # Only check string columns
        pattern = r'https?://|www\.'
        matches = df[col_name].astype(str).str.contains(pattern, regex=True, na=False)
        return matches.any()
    return False

def should_delete_column(df, col_name):
    """Determine if a column should be deleted based on criteria"""
    # Check if it's empty
    if is_empty_column(df, col_name):
        return True
    
    # Check if it contains links but is not a protected column name
    protected_columns = ['caption', 'title', 'description']
    column_lower = col_name.lower() if isinstance(col_name, str) else str(col_name).lower()
    
    if column_lower in protected_columns:
        return False  # Never delete these columns even if they contain links
        
    if contains_links(df, col_name):
        return True  # Delete other columns with links
        
    return False  # Keep other columns

def is_sheet_empty(df):
    """Check if all columns in the sheet are empty"""
    return all(is_empty_column(df, col) for col in df.columns)

def get_platform_files(directory):
    """Get a list of Excel files in the specified directory"""
    files = []
    if os.path.exists(directory):
        for filename in os.listdir(directory):
            if filename.endswith('.xlsx'):
                files.append(os.path.join(directory, filename))
    return files

def get_sheets_from_file(file_path):
    """Get a list of sheets from an Excel file"""
    try:
        wb = load_workbook(file_path, read_only=True)
        return wb.sheetnames
    except Exception as e:
        print(f"Error reading {file_path}: {str(e)}")
        return []

def print_sheet_list(sheets, with_numbers=True):
    """Print a list of sheets with optional numbering"""
    if with_numbers:
        for i, sheet in enumerate(sheets, 1):
            print(f"{i}. {sheet}")
    else:
        for sheet in sheets:
            print(f"- {sheet}")

def select_sheet_with_prompt(sheets, prompt_message):
    """Prompt user to select a sheet from a list"""
    print_sheet_list(sheets)
    
    while True:
        try:
            selection = input(prompt_message)
            if selection.strip().lower() == 'skip':
                return None
            idx = int(selection) - 1
            if 0 <= idx < len(sheets):
                return sheets[idx]
            else:
                print(f"Please enter a number between 1 and {len(sheets)}")
        except ValueError:
            print("Please enter a valid number or 'skip'")

def create_sheet_mapping(platform_dirs):
    """Create a mapping of sheet names across platforms based on user input"""
    # Get a sample file from each platform
    platform_samples = {}
    platform_sheets = {}
    
    for platform, directory in platform_dirs.items():
        files = get_platform_files(directory)
        if files:
            platform_samples[platform] = files[0]
            platform_sheets[platform] = get_sheets_from_file(files[0])
        else:
            print(f"No Excel files found for {platform}")
    
    if len(platform_samples) < 2:
        print("Need at least two platforms with files to create mappings")
        return None
    
    # Start with YouTube as the reference if available
    reference_platform = 'youtube' if 'youtube' in platform_sheets else list(platform_sheets.keys())[0]
    reference_sheets = platform_sheets[reference_platform]
    
    if not reference_sheets:
        print(f"No sheets found in {reference_platform} files")
        return None
    
    print(f"\n==== Sheet Mapping Configuration ====")
    print(f"Using {reference_platform.upper()} as the reference platform")
    print(f"Available sheets in {reference_platform.upper()}:")
    print_sheet_list(reference_sheets, with_numbers=False)
    
    # Create mapping
    mapping = {}
    
    for platform in platform_sheets:
        if platform == reference_platform:
            continue
            
        print(f"\nMapping {platform.upper()} sheets to {reference_platform.upper()} sheets:")
        platform_mapping = {}
        
        for ref_sheet in reference_sheets:
            print(f"\nFor {reference_platform.upper()} sheet '{ref_sheet}', select the equivalent in {platform.upper()}:")
            
            if platform_sheets[platform]:
                equivalent = select_sheet_with_prompt(
                    platform_sheets[platform],
                    f"Enter number (or 'skip' if no equivalent): "
                )
                
                if equivalent:
                    platform_mapping[equivalent] = ref_sheet
            else:
                print(f"No sheets available in {platform}")
        
        mapping[platform] = platform_mapping
    
    return reference_platform, mapping

def clean_excel_files(input_dirs, sheet_mapping=None):
    """Clean Excel files in the specified directories with optional sheet mapping"""
    for platform, directory in input_dirs.items():
        if not os.path.exists(directory):
            print(f"Directory not found: {directory}")
            continue
            
        for filename in os.listdir(directory):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(directory, filename)
                print(f"Processing file: {file_path}")
                
                try:
                    # First step: Handle sheet operations using openpyxl
                    wb = load_workbook(file_path)
                    
                    # Track sheets to delete and rename
                    sheets_to_delete = []
                    sheets_to_rename = {}
                    
                    for sheet_name in wb.sheetnames:
                        # Apply user-defined sheet mapping if available
                        if sheet_mapping and platform in sheet_mapping and sheet_name in sheet_mapping[platform]:
                            target_name = sheet_mapping[platform][sheet_name]
                            sheets_to_rename[sheet_name] = target_name
                            continue
                            
                        # Delete sheets with "hashtag" or "hashtags"
                        if re.search(r'hashtags?', sheet_name, re.IGNORECASE):
                            sheets_to_delete.append(sheet_name)
                        
                        # Rename sheets by removing "Post" or "Post Information"
                        elif re.search(r'Post Information|Post', sheet_name, re.IGNORECASE):
                            new_name = re.sub(r'Post Information|Post', '', sheet_name, flags=re.IGNORECASE).strip()
                            if new_name and new_name != sheet_name:
                                sheets_to_rename[sheet_name] = new_name
                    
                    # Delete sheets
                    for sheet_name in sheets_to_delete:
                        print(f"  Deleting sheet: {sheet_name}")
                        del wb[sheet_name]
                    
                    # Rename sheets
                    for old_name, new_name in sheets_to_rename.items():
                        if new_name in wb.sheetnames:
                            print(f"  Cannot rename {old_name} to {new_name} (already exists)")
                            continue
                        print(f"  Renaming sheet: {old_name} â†’ {new_name}")
                        wb[old_name].title = new_name
                    
                    # Save changes from sheet operations
                    wb.save(file_path)
                    
                    # Second step: Handle column operations and check for empty sheets
                    xl = pd.ExcelFile(file_path)
                    
                    # Get a list of empty sheets to delete
                    empty_sheets = []
                    sheet_data = {}
                    
                    for sheet_name in xl.sheet_names:
                        df = xl.parse(sheet_name)
                        
                        if df.empty or is_sheet_empty(df):
                            empty_sheets.append(sheet_name)
                            print(f"  Sheet {sheet_name} is empty and will be deleted")
                            continue
                        
                        # Find columns to delete based on updated criteria
                        cols_to_delete = []
                        for col in df.columns:
                            if should_delete_column(df, col):
                                cols_to_delete.append(col)
                        
                        # Remove identified columns
                        if cols_to_delete:
                            print(f"  Removing {len(cols_to_delete)} columns from sheet {sheet_name}")
                            df = df.drop(columns=cols_to_delete)
                        
                        # Check if sheet becomes empty after column deletions
                        if df.empty or len(df.columns) == 0:
                            empty_sheets.append(sheet_name)
                            print(f"  Sheet {sheet_name} has no data after column deletions and will be deleted")
                        else:
                            sheet_data[sheet_name] = df
                    
                    # Delete empty sheets
                    wb = load_workbook(file_path)
                    for sheet_name in empty_sheets:
                        if sheet_name in wb.sheetnames:
                            del wb[sheet_name]
                    
                    # Check if we have at least one sheet left
                    if len(sheet_data) == 0 and len(wb.sheetnames) == 0:
                        print(f"  Warning: All sheets in {file_path} would be deleted. Creating a dummy sheet.")
                        # Create a dummy sheet if all sheets would be deleted
                        ws = wb.create_sheet("Empty_Workbook")
                        ws['A1'] = "This workbook had no valid data."
                    
                    wb.save(file_path)
                    
                    # Save non-empty sheets with processed data
                    if sheet_data:
                        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                            for sheet_name, df in sheet_data.items():
                                df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    print(f"Completed processing: {file_path}")
                    
                except Exception as e:
                    print(f"Error processing {file_path}: {str(e)}")

def main():
    """Main function to run the script with CLI interface"""
    input_dirs = {
        'facebook': r"C:\Users\Admin\Desktop\Final\data\input\facebook",
        'instagram': r"C:\Users\Admin\Desktop\Final\data\input\instagram",
        'youtube': r"C:\Users\Admin\Desktop\Final\data\input\youtube"
    }
    
    print("=== Excel File Sheet Standardization Tool ===")
    print("This tool helps you standardize sheet names across platforms.")
    
    while True:
        print("\nSelect an option:")
        print("1. Auto-clean files (using existing logic)")
        print("2. Create sheet mapping and clean files")
        print("3. Exit")
        
        choice = input("Enter your choice (1-3): ")
        
        if choice == '1':
            print("\nRunning auto-clean without sheet mapping...")
            clean_excel_files(input_dirs)
            print("Excel cleaning process completed!")
            
        elif choice == '2':
            print("\nCreating sheet mapping across platforms...")
            mapping_result = create_sheet_mapping(input_dirs)
            
            if mapping_result:
                reference_platform, platform_mapping = mapping_result
                
                print("\nSheet mapping summary:")
                for platform, mapping in platform_mapping.items():
                    print(f"\n{platform.upper()} to {reference_platform.upper()} mappings:")
                    if mapping:
                        for source, target in mapping.items():
                            print(f"  '{source}' will be renamed to '{target}'")
                    else:
                        print("  No mappings defined")
                
                confirm = input("\nProceed with cleaning using this mapping? (y/n): ")
                if confirm.lower() == 'y':
                    clean_excel_files(input_dirs, platform_mapping)
                    print("Excel cleaning process with sheet mapping completed!")
                else:
                    print("Operation cancelled.")
            else:
                print("Sheet mapping creation failed or was cancelled.")
                
        elif choice == '3':
            print("Exiting program.")
            break
            
        else:
            print("Invalid choice. Please enter 1, 2, or 3.")

if __name__ == "__main__":
    main()